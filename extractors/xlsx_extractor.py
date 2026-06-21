from openpyxl import load_workbook

def extract_text(path):
    wb = load_workbook(path, read_only=True, data_only=True)

    full_text = []

    for sheet in wb.worksheets:
        full_text.append(f"Sheet: {sheet.title}")

        for row in sheet.iter_rows(values_only=True):
            row_text = [str(cell) for cell in row if cell is not None]

            if row_text:
                full_text.append(" | ".join(row_text))

        full_text.append("")

    return "\n".join(full_text)